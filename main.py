from openpyxl import Workbook, load_workbook
from openpyxl.utils import rows_from_range, get_column_letter
from config import HEADERS, ROW_START, ROW_ITEM_COUNT, ROW_GAP, VAT_POSITIONS
from tkinter import *
from tkinter import filedialog, messagebox
import os
import time


# TODO: Cache the result of merged_cells
# Returns the value for a cell. If multiple cells are merged,
# the merge range will have the same value
def get_value_with_merge_lookup(sheet, cell, merged_cells_ranges):
    # EG: A12, B12
    idx = cell.coordinate

    for range_ in merged_cells_ranges:
        merged_cells = list(rows_from_range(str(range_)))
        for row in merged_cells:
            if idx in row:
                return str(sheet[merged_cells[0][0]].value)
    return str(sheet[idx].value)


# Returns the value that will be added in the new worksheet
# Based on the values extracted from the original
def get_cell_value_for_new_worksheet(header_item, header, values, vat,
                                     line_number):
    value = None
    if header_item['value'] is not None:
        value = header_item['value']
    else:
        if header_item['position_original_document'] is not None:
            # Alter the data if needed
            returned_value = values[header_item['position_original_document']]
            if header == 'Numar':
                returned_value = returned_value.replace('/', '')
            if header == 'Cod fiscal':
                returned_value = returned_value.replace('RO', '')
            value = returned_value
        # Special scenario for VAT specific fields
        else:
            if header == 'Numar linie':
                value = line_number
            elif header == 'Valoare neta totala' or header == 'Pret de lista' or header == 'Valoare fara tva':
                value = vat['total_net_amount']
            elif header == 'Valoare TVA' or header == 'Val TVA':
                value = vat['vat']
            elif header == 'Total document' or header == 'Val cu TVA':
                value = vat['total_amount']
            elif header == 'Cod articol':
                value = vat['article_code']
            elif header == 'Denumire articol':
                value = vat['article_description']
            elif header == 'Cota TVA':
                value = vat['vat_percent']
    return value


# Do not consider invoices with a total ammount of 0
def invoice_should_be_processed(values):
    return float(values[12]) != 0


# Get the VATs that are in the curent row
def get_vat_rates(values):
    return [{
        'total_net_amount':
        values[vat['total_amount']],
        'vat':
        values[vat['vat']],
        'total_amount':
        "{:.2f}".format(
            float(values[vat['total_amount']]) + float(values[vat['vat']])),
        'article_code':
        vat['article_code'],
        'article_description':
        vat['article_description'],
        'vat_percent':
        vat['vat_percent']
    } for vat in VAT_POSITIONS.values()
            if float(values[vat['total_amount']]) > 0.00]


# If the first cell is None, there are no more entries in the page
def check_page_finished(values):
    return values[0] == 'None' or not values[0].isdigit()


def convert_xls_to_xlsx(filename):
    # /cacat/pisat.xls
    destination = '/'.join(filename.split('/')[0:-1])
    sys = os.system(
        f'/Applications/LibreOffice.app/Contents/MacOS/soffice --convert-to xlsx "{filename}" --outdir "{destination}"'
    )
    time.sleep(3)


def convert_xlxs_to_xls(filename):
    destination = '/'.join(filename.split('/')[0:-1])
    sys = os.system(
        f'/Applications/LibreOffice.app/Contents/MacOS/soffice --convert-to xls "{filename}" --outdir "{destination}"'
    )
    time.sleep(3)


def main_app(filename):
    convert_xls_to_xlsx(filename)
    # Original document
    wb = load_workbook(filename.replace('.xls', '.xlsx'))
    ws = wb.active
    merged_cells_ranges = ws.merged_cells.ranges

    # New document
    new_wb = Workbook()
    new_ws = new_wb.active

    # Start processing
    ok = True
    row_new_document = 1
    row_start = ROW_START
    row_end = ROW_START + ROW_ITEM_COUNT
    # Iterate over rows in the original document
    row_original_document = row_start
    max_rows_original_document = ws.max_row
    while row_original_document <= max_rows_original_document:
        values = []
        # Going through each cell of a row in the original document
        # Get cell values from the original document after adjusting the merged cells
        for cell in ws[row_original_document]:
            values.append(
                get_value_with_merge_lookup(ws, cell, merged_cells_ranges))

        if check_page_finished(values) == True:
            row_original_document += ROW_GAP - 1
            continue

        print(f"Processing row {row_original_document}")
        # If total ammount is 0, skip this row
        if not invoice_should_be_processed(values):
            row_original_document += 1
            continue

        # Detect the different rates for VAT
        vats = get_vat_rates(values)

        line_number = 1
        for vat in vats:
            # Start building the new file based on the specified headers
            for col, header in enumerate(HEADERS):
                header_item = HEADERS[header]
                value = get_cell_value_for_new_worksheet(
                    header_item, header, values, vat, line_number)
                char_new_ws = get_column_letter(col + 1)
                new_ws[char_new_ws + str(row_new_document)] = value
            # Process a new row in the new document
            row_new_document += 1
            line_number += 1
        row_original_document += 1

    new_filename = '/'.join(
        filename.split('/')[0:-1]) + '/' + filename.split('/')[-1].replace(
            '.xls', '_nou.xlsx')
    new_wb.save(new_filename)
    time.sleep(5)
    wb.close()
    convert_xlxs_to_xls(new_filename)
    os.remove(filename.replace('.xls', '.xlsx'))
    os.remove(new_filename)
    messagebox.showinfo("Atentie", "Fisierul a fost procesat si salvat")


def browseFiles():
    filename = filedialog.askopenfilename(initialdir="/",
                                          title="Select a File",
                                          filetypes=(("Text files", "*.txt*"),
                                                     ("all files", "*.*")))

    # Change label contents
    # label_file_explorer.configure(
    #     text="Fisierul a fost deschis si este procesat")
    main_app(filename=filename)


# Create the root window
window = Tk()

# Set window title
window.title('AlyContab')

# Set window size
window.geometry("500x500")
window.resizable(False, False)

#Set window background color
window.config(background="white")

# Create a File Explorer label
label_file_explorer = Label(
    window, text="Selecteaza un fisier pentru a incepe procesul")

button_explore = Button(window, text="Adauga fisier", command=browseFiles)

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns
label_file_explorer.grid(column=1, row=1)

button_explore.grid(column=1, row=2)

# Let the window wait for any events
window.mainloop()